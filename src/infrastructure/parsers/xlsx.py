from __future__ import annotations

import hashlib
from io import BytesIO

from openpyxl import load_workbook

from domain.models import ParsedDocument, ParsedElement


class XLSXParser:
    name, version = "openpyxl", "1.0.0"

    def supports(self, file_type: str) -> bool: return file_type.lower() == "xlsx"

    def parse(self, data: bytes, document_name: str) -> ParsedDocument:
        try: workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
        except Exception as exc: raise ValueError("Corrupted XLSX") from exc
        elements, order = [], 0
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                while values and values[-1] == "": values.pop()
                if any(values): rows.append(values)
            if not rows: continue
            max_width = max(map(len, rows)); rows = [row + [""] * (max_width - len(row)) for row in rows]
            text = "\n".join(" | ".join(row) for row in rows)
            elements.append(ParsedElement(element_type="table", text=text, order=order, heading_path=[sheet.title], table_id=f"sheet:{sheet.title}", table_rows=rows, source_ref=f"sheet:{sheet.title}!A1")); order += 1
        return ParsedDocument(document_id=hashlib.sha256(document_name.encode()).hexdigest()[:16], document_name=document_name,
            file_type="xlsx", checksum=hashlib.sha256(data).hexdigest(), parser_name=self.name, parser_version=self.version,
            elements=elements, metadata={"sheet_names": workbook.sheetnames, "table_count": len(elements)})
